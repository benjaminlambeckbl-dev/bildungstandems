# -*- coding: utf-8 -*-
"""Erzeugt docs/Berechtigungsmatrix.xlsx (Workshop-Stand 11.06.2026).
4 Rollen (Admin/Trainer/Koordination/Schüler-Coach), Ausfüll-/Bestätigungsvorlage."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1D1F4E"; NAVY_LT="E9EAF3"
F={
 "v":("Verwalten","70AD79","FFFFFF"),   # voll: anlegen/bearbeiten/löschen
 "e":("Beitragen","C6EFCE","1C2230"),    # erstellen/schreiben im eigenen Bereich
 "l":("Ansehen","DDEBF7","1C2230"),      # nur lesen/nutzen
 "x":("—","F2F2F2","8A97A5"),            # kein Zugriff
}
thin=Side(style="thin",color="D9DEE6"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
ROLES=["Admin (ZSB)","Trainer:in","Koordination (Lehrkraft)","Schüler:in (Coach)"]

# (Sektion, [(Feature, [4 Codes], Hinweis)]); Code = 'v'/'e'/'l'/'x' oder ('e','Scope')
SEC=[
 ("1 · Material & Inhalte",[
   ("Material ansehen / suchen",["v",("l","betreute"),"l",("l","Teilmenge")],"Suche über Kategorien/Tags (keine PDF-Volltextsuche)"),
   ("Material hochladen & taggen",["v","x","x","x"],"Nur Admin – kuratierte Qualität"),
   ("Sichtbarkeit/Freigabe festlegen",["v","x","x","x"],"„alle“ oder „nur Lehrkräfte“"),
   ("Zeitgesteuerte Freischaltung",["v","x","x","x"],"z. B. Übergangsmaterial erst im letzten Quartal"),
   ("Kategorien & Tags pflegen",["v","x","x","x"],"feste Kategorien, zentral"),
 ]),
 ("2 · Termine",[
   ("Globale Termine (Abschlussfeier, Fortbildung)",["v","x","x","x"],"Admin → alle; Empfänger nur zu-/absagen"),
   ("Trainingstermine",["v",("e","betreute"),"l","x"],"Trainer:in für betreute Schulen (~3/Jahr)"),
   ("Tandem-Termine",["v","x",("e","eigene Schule"),"x"],"Koordination; Partnerschule sieht automatisch"),
   ("Termine ansehen",[("v","alle"),("l","betreute"),("l","Schule"),("l","eigene")],""),
   ("Teilnahme zu-/absagen",["e","e","e",("e","eigene")],""),
   ("Warnung fehlende Termine",["v",("l","betreute"),"x","x"],"Admin-Dashboard / Ampel gelb"),
 ]),
 ("3 · Kommunikation (Jugendschutz)",[
   ("Event-Nachrichten erhalten",["l","l","l","l"],"Termin abgesagt/verschoben/neu – an eigene Termine gekoppelt"),
   ("Direktnachricht Koord ↔ Trainer",["v","e","e","x"],"beide Richtungen"),
   ("Direktnachricht Koord ↔ Koord",["v","x","e","x"],"nur Koordinationen einer Tandem-Schule"),
   ("Direktnachricht Trainer ↔ Admin",["e","e","x","x"],""),
   ("Schüler-Chat",["x","x","x","x"],"BEWUSST NICHT vorgesehen (Jugendschutz)"),
   ("Gruppenchat Tandem-Teams",["v","x","e","x"],"OFFEN – wg. Jugendschutz zu klären"),
 ]),
 ("4 · Feedback & Monitoring",[
   ("Termin-Feedback (Smiley)",["x","e","e","e"],"nach Training/Tandem-Treffen; bei „schlecht“ Textfeld"),
   ("Reflexion schreiben",["x","x","e","x"],"Koordination, optional"),
   ("Ampel-Umfrage (alle 6 Wochen)",["x","x","e","x"],"„Wie läuft das Projekt?“ grün/gelb/rot"),
   ("„Hilfe nötig?“ / SOS auslösen",["x","x","e","e"],"Schüler (nicht-akut) + Koordination → ZSB"),
   ("Monitoring/Ampel-Übersicht",[("v","alle"),("l","betreute"),("l","Schule"),"x"],""),
   ("Auswertung / Diagramme",[("v","alle"),("l","betreute"),("l","Schule"),"x"],""),
 ]),
 ("5 · Export & Zertifikate",[
   ("Daten-/Excel-Export",["v","x","x","x"],"Statistiken, Stadtverteilung, Stimmung, Namen"),
   ("Zertifikat erstellen/abrufen",["v","x",("e","eigene SuS"),("l","eigenes")],""),
 ]),
 ("6 · Nutzer-, Schul- & Jahresverwaltung",[
   ("Schulen anlegen/bearbeiten",["v","x","x","x"],""),
   ("Schul-Kooperation zuordnen (1:n)",["v","x","x","x"],"1 weiterführende ↔ bis 3 Grundschulen"),
   ("Koordinator:innen/Trainer:innen anlegen",["v","x","x","x"],"+ Login-Daten"),
   ("Trainer ↔ Schule zuweisen/auflösen",["v","x","x","x"],"Trainer:in kann nicht selbst lösen"),
   ("Schüler-Registrierung freigeben/sperren",["v","x",("e","eigene Schule"),"x"],"QR-Anmeldung, Koordination bestätigt einzeln"),
   ("Schüler entfernen/korrigieren",["v","x",("e","eigene"),"x"],"mit Verlauf/Historie (kein Überschreiben)"),
   ("Schuljahr / Jahreswechsel verwalten",["v","x","x","x"],"Enddatum, Auto-Löschung SuS am Jahresende"),
 ]),
]

wb=Workbook(); ws=wb.active; ws.title="Berechtigungsmatrix"; ws.sheet_view.showGridLines=False
def put(r,c,v,*,bold=False,fill=None,fc="1C2230",size=10,align="left",wrap=False,border=False):
    cl=ws.cell(r,c,v); cl.font=Font(name="Calibri",size=size,bold=bold,color=fc)
    cl.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if fill: cl.fill=PatternFill("solid",fgColor=fill)
    if border: cl.border=BORDER
    return cl

put(1,1,"Berechtigungsmatrix – BildungsTandems / ZSB",bold=True,size=15,fc=NAVY)
put(2,1,"Stand Workshop 11.06.2026 · Vorschlag zum Bestätigen/Anpassen durch die ZSB · Coachingspace GmbH",size=9,fc="5C6B7A")
put(3,1,"Legende:",bold=True,size=9,fc=NAVY)
for i,k in enumerate(["v","e","l","x"]):
    t,bg,fc=F[k]; put(3,2+i,t,fill=bg,fc=fc,align="center",size=9,bold=True,border=True)
put(3,6,"Geltungsbereich in Klammern: alle / betreute Schulen / Schule / eigene",size=9,fc="5C6B7A")
put(4,1,"Doppelrollen möglich (z. B. Admin + Trainer:in) – Rechte addieren sich.",size=9,fc="5C6B7A")

HEAD=6
put(HEAD,1,"Feature",bold=True,fill=NAVY,fc="FFFFFF",border=True)
for j,role in enumerate(ROLES): put(HEAD,2+j,role,bold=True,fill=NAVY,fc="FFFFFF",align="center",wrap=True,border=True)
put(HEAD,6,"Hinweis",bold=True,fill=NAVY,fc="FFFFFF",border=True)

r=HEAD+1
for sec,rows in SEC:
    put(r,1,sec,bold=True,fill=NAVY_LT,fc=NAVY,border=True)
    for j in range(1,6): put(r,1+j,"",fill=NAVY_LT,border=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); put(r,1,sec,bold=True,fill=NAVY_LT,fc=NAVY,border=True)
    r+=1
    for feat,codes,hint in rows:
        put(r,1,feat,bold=True,wrap=True,border=True)
        for j,spec in enumerate(codes):
            key,scope=(spec if isinstance(spec,tuple) else (spec,None))
            t,bg,fc=F[key]; disp=t if not scope else f"{t}\n({scope})"
            put(r,2+j,disp,fill=bg,fc=fc,align="center",wrap=True,border=True,bold=(key=="v"))
        put(r,6,hint,size=8,fc="5C6B7A",wrap=True,border=True)
        r+=1

ws.column_dimensions["A"].width=40
for j in range(len(ROLES)): ws.column_dimensions[get_column_letter(2+j)].width=18
ws.column_dimensions["F"].width=40
ws.freeze_panes="B7"; ws.row_dimensions[HEAD].height=30

# ---- Blatt 2: Direktchat ----
ws2=wb.create_sheet("Direktchat"); ws2.sheet_view.showGridLines=False
def c2(r,c,v,*,bold=False,fill=None,fc="1C2230",size=10,align="center",wrap=False,border=False):
    cl=ws2.cell(r,c,v); cl.font=Font(name="Calibri",size=size,bold=bold,color=fc)
    cl.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if fill: cl.fill=PatternFill("solid",fgColor=fill)
    if border: cl.border=BORDER
c2(1,1,"Direktchat / Kommunikation – wer darf wen anschreiben?",bold=True,size=14,fc=NAVY,align="left")
c2(2,1,"ja = erlaubt · — = nicht vorgesehen · Schüler:innen: KEIN Chat (nur Event-Nachrichten + SOS)",size=9,fc="5C6B7A",align="left")
GREEN=("ja","C6EFCE","1C2230"); GREY=("—","F2F2F2","8A97A5"); SELF=("–","E9EAF3","8A97A5")
roles_s=["Admin","Trainer","Koordination","Schüler:in"]; H=4
c2(H,1,"Initiator \\ Empfänger",bold=True,fill=NAVY,fc="FFFFFF",align="left",border=True,wrap=True)
for j,role in enumerate(roles_s): c2(H,2+j,role,bold=True,fill=NAVY,fc="FFFFFF",border=True,wrap=True)
chat={
 "Admin":["self","ja","ja","nein"],
 "Trainer":["ja","ja","ja","nein"],
 "Koordination":["ja","ja","ja*","nein"],
 "Schüler:in":["nein","nein","nein","nein"],
}
M={"ja":GREEN,"ja*":GREEN,"nein":GREY,"self":SELF}
for i,role in enumerate(roles_s):
    rr=H+1+i; c2(rr,1,role,bold=True,fill=NAVY_LT,fc=NAVY,align="left",border=True)
    for j,code in enumerate(chat[role]):
        disp,bg,fc=M[code];
        if code=="ja*": disp="ja*"
        c2(rr,2+j,disp,fill=bg,fc=fc,border=True)
c2(H+6,1,"*  Koordination ↔ Koordination nur zwischen den Koordinationen einer Tandem-Schule.",size=8,fc="5C6B7A",align="left")
c2(H+7,1,"Statt freiem Chat: event-basierte Vorlage-Nachrichten (Termin abgesagt/verschoben/neu).",size=8,fc="5C6B7A",align="left")
ws2.column_dimensions["A"].width=24
for j in range(len(roles_s)): ws2.column_dimensions[get_column_letter(2+j)].width=16
ws2.row_dimensions[H].height=26

wb.save("docs/Berechtigungsmatrix.xlsx"); print("ok")
