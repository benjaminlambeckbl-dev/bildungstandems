# -*- coding: utf-8 -*-
"""Erzeugt docs/Feature-Liste-Aufwand.xlsx – ausdifferenziert, mit Stundenaufwand,
Pflicht/Optional und interaktivem Auswahl-Konfigurator (SUMIF)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY="1D1F4E"; NAVY_LT="E9EAF3"; ORANGE="F59E1B"
PFILL="1D1F4E"; OFILL="DDEBF7"
thin=Side(style="thin",color="D9DEE6")
B=Border(left=thin,right=thin,top=thin,bottom=thin)
RATE=100  # €/h  (= 800 €/PT)

# (Kategorie, [(Feature, Beschreibung, Typ 'P'/'O', Stunden)])
DATA=[
 ("0 · Pflicht – Fundament (Voraussetzung für alles)", [
   ("Backend-Setup","Projektgerüst Java 25 / Spring Boot, Struktur, Build/CI.","P",8),
   ("Datenbank-Schema + Migrationen","PostgreSQL-Datenmodell aller Entitäten + Versionierung.","P",16),
   ("Hosting Hetzner","Server (Nürnberg/Falkenstein), Deploy-Pipeline, Domain, HTTPS, Backups.","P",16),
   ("DSGVO-/Betriebs-Grundsetup","AVV-Setup, Logging, Monitoring, Datensparsamkeit.","P",8),
   ("API-Grundgerüst + Frontend-Anbindung","REST-API; bestehendes Angular-Frontend gegen echtes Backend.","P",16),
   ("Login & Authentifizierung","Anmeldung, Session/JWT, Passwort-Reset.","P",16),
   ("Rollen-/Rechte-Framework","4 Rollen (Admin/Trainer/Koordination/Coach) + Doppelrollen, serverseitig durchgesetzt.","P",20),
   ("Schul- & Kooperations-Datenmodell","Weiterführende ↔ bis 3 Grundschulen (1:n), Tandem-Struktur.","P",8),
   ("Schüler-Registrierung (QR)","QR je Schule, Self-Registrierung (nur Name), Koordinator-Freigabe, Sperre.","P",24),
   ("Admin-Verwaltung Basis","Schulen anlegen; Koordinator:innen/Trainer:innen anlegen + Schul-Zuordnung.","P",16),
   ("Tests + CI","Unit-/Integrationstests, Continuous Integration.","P",16),
   ("Abnahme, Doku, Übergabe","Abnahmetests, Kurz-Doku, Einweisung.","P",12),
 ]),
 ("A · Terminverwaltung", [
   ("Termin-Modell (3 Typen)","Global / Training / Tandem als Datenmodell + Logik.","O",12),
   ("Termine anlegen/bearbeiten/absagen","Rollenabhängig (Admin global, Trainer Training, Koordination Tandem).","O",12),
   ("Termin-Detail + Zu-/Absage","Detailseite, Teilnahme zu-/absagen.","O",8),
   ("Kalenderansicht","Monatskalender (persönlich/schul/global).","O",6),
   ("Erinnerungen vor Termin","Hinweis vor Treffen (benötigt Push oder E-Mail – Modul F).","O",6),
   ("Warnung fehlende Trainingstermine","Admin-Dashboard hebt Schulen ohne Termine hervor.","O",8),
   ("ICS-Export","Termine in eigenen Kalender exportieren.","O",6),
   ("Terminabstimmung (Poll)","Abstimmen über mögliche Termine.","O",12),
 ]),
 ("B · Materialbereitstellung", [
   ("Material-Modell + Kategorien/Tags","Feste Kategorien (Methodenkarten, Begleitmaterial …) + Tags.","O",8),
   ("Material ansehen + Suche","Liste, Suche nach Kategorie/Tags (keine PDF-Volltextsuche).","O",8),
   ("Datei-Upload / Storage","PDF/Bilder hochladen (nur Admin), Rechte.","O",12),
   ("Web-Ansicht + CMS-Editor","Inhalte direkt lesbar; Rich-Text-Editor (fett, Überschriften).","O",24),
   ("Zeitgesteuerte Freischaltung","Inhalte ab Datum/Phase sichtbar.","O",10),
   ("Auto-Freischaltung nach Training","Trigger: nach 1./2. Training weiteres Material.","O",8),
   ("Nutzungs-Statistik","Aufruf-Zähler je Material (aggregiert).","O",8),
 ]),
 ("C · Lernen", [
   ("Lernpfade","Schritte, Quiz, Vorlesen, Fortschritt.","O",14),
   ("FAQ-Bereich","Durchsuchbar, kategorisiert.","O",6),
 ]),
 ("D · Kommunikation (Jugendschutz-konform)", [
   ("Event-Nachrichten","Vorlagen: Termin abgesagt/verschoben/neu (kein Freitext-Chat für SuS).","O",10),
   ("Direktnachrichten (Erwachsene)","Koord↔Trainer, Koord↔Koord (Tandem-Schule), Trainer↔Admin.","O",16),
   ("Gruppenchat Tandem-Teams","Koordination legt Gruppe für Coach-Teams an.","O",14),
 ]),
 ("E · Feedback & Monitoring", [
   ("Termin-Feedback","Smiley (gut/mittel/schlecht), bei schlecht Textfeld.","O",8),
   ("Reflexion (Koordination)","Optionale ausführlichere Reflexion.","O",6),
   ("Ampel-Umfrage (6 Wochen)","„Wie läuft das Projekt?“ grün/gelb/rot je Schule.","O",12),
   ("„Hilfe nötig?“ / SOS","Nicht-akute Hilfeanfrage → Push an Koordination.","O",8),
   ("Monitoring-Dashboard","KPIs (Kinder, Tandems) + Ampel-Farbübersicht aller Schulen.","O",16),
   ("Änderungs-/Drop-out-Historie","Protokoll statt Überschreiben (z. B. Gruppe 15→5).","O",10),
 ]),
 ("F · Benachrichtigungen", [
   ("Push-Notifications","DSGVO-konforme Lösung (EU/Open-Source) – Recherche + Umsetzung. RISIKO/offen.","O",28),
   ("E-Mail-Versand","Erinnerungen/Einladungen per E-Mail.","O",8),
 ]),
 ("G · Export & Zertifikate", [
   ("Coach-Zertifikat (PDF)","Druckbare Urkunde inkl. Namenserfassung.","O",8),
   ("Excel-/CSV-Export","Namen, Trainingsanzahl, Stadtverteilung, Schülerzahlen.","O",12),
   ("Stimmungs-Export","Feedback-/Stimmungsverläufe je Schule exportierbar.","O",6),
 ]),
 ("H · Verwaltung erweitert", [
   ("Schul-Filter/Suche","Filter nach Stadt, Schulform, Sozialindex (>100 Schulen).","O",10),
   ("Schuljahres-Logik","Enddatum manuell, Jahre deaktivierbar, Auto-Löschung SuS am Jahresende.","O",16),
   ("Teilnahme-Umfrage + Pause/Resume","Jahresend-Abfrage, Schulen pausieren statt löschen.","O",10),
   ("Zeit-Trigger","z. B. „2 Wochen nach Schulstart Erinnerung“.","O",10),
 ]),
 ("I · Design & UX", [
   ("Schlankes Schüler-UI","Minimalistischeres Redesign, weniger Emojis/Scrollen.","O",12),
   ("Responsive-Feinschliff","Optimierung Handy/Desktop.","O",6),
   ("CI-/Theming-Anpassbarkeit","Farben/Design anpassbar.","O",8),
 ]),
]

wb=Workbook(); ws=wb.active; ws.title="Feature-Liste"
ws.sheet_view.showGridLines=False
def put(r,c,v,*,bold=False,fill=None,fc="1C2230",size=10,align="left",wrap=False,border=False,numfmt=None):
    cl=ws.cell(r,c,v); cl.font=Font(name="Calibri",size=size,bold=bold,color=fc)
    cl.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if fill: cl.fill=PatternFill("solid",fgColor=fill)
    if border: cl.border=B
    if numfmt: cl.number_format=numfmt
    return cl

# Spalten: A Kategorie/Feature, B Beschreibung, C Typ, D Aufwand(h), E Auswahl(x)
put(1,1,"Feature-Liste mit Aufwand – BildungsTandems / ZSB",bold=True,size=15,fc=NAVY)
put(2,1,"Pflicht (Fundament) ist Voraussetzung. Optionale Bausteine in Spalte „Auswahl“ mit  x  wählen – Stunden & Kosten rechnen automatisch.",size=9,fc="5C6B7A")

# Datenbereich vorab bestimmen
HEAD=8
first=HEAD+1
# zähle Datenzeilen inkl. Sektionsköpfe
rowcount=0
for cat,rows in DATA: rowcount+=1+len(rows)
last=first+rowcount-1
Dcol="D"; Ecol="E"
sumif=f'=SUMIF(${Ecol}${first}:${Ecol}${last},"x",${Dcol}${first}:${Dcol}${last})'
pflicht_h=sum(h for _,rows in DATA for _,_,t,h in rows if t=="P")
opt_h=sum(h for _,rows in DATA for _,_,t,h in rows if t=="O")

# Summary-Block (oben rechts)
put(1,4,"Stundensatz",bold=True,fc=NAVY,align="right"); put(1,5,RATE,align="right",border=True,numfmt='#,##0 "€/h"')
put(2,4,"Pflicht (h)",bold=True,fc=NAVY,align="right"); put(2,5,pflicht_h,align="right",border=True,numfmt='#,##0 "h"')
put(3,4,"Gewählt gesamt (h)",bold=True,fc=NAVY,align="right"); put(3,5,sumif,align="right",border=True,numfmt='#,##0 "h"',fill="FFF6E5")
put(4,4,"Geschätzte Kosten",bold=True,fc=NAVY,align="right"); put(4,5,f"=E3*E1",align="right",border=True,numfmt='#,##0 "€"',fill="FFF6E5")
put(5,4,"Budget ZSB",bold=True,fc=NAVY,align="right"); put(5,5,20000,align="right",border=True,numfmt='#,##0 "€"')
put(6,4,"Differenz",bold=True,fc=NAVY,align="right"); put(6,5,"=E5-E4",align="right",border=True,numfmt='#,##0 "€"')
put(4,1,"Legende:",bold=True,size=9,fc=NAVY)
put(4,2,"Pflicht",fill=PFILL,fc="FFFFFF",bold=True,align="center",size=9,border=True)
put(5,2,"Optional",fill=OFILL,align="center",size=9,border=True)
put(6,1,f"Alles optional zusätzlich: {opt_h} h (alle Module).",size=9,fc="5C6B7A")

# Kopfzeile
heads=[("Kategorie / Feature",34),("Beschreibung",60),("Typ",11),("Aufwand (h)",12),("Auswahl",10)]
for j,(h,w) in enumerate(heads):
    put(HEAD,1+j,h,bold=True,fill=NAVY,fc="FFFFFF",align="left" if j<2 else "center",border=True)
    ws.column_dimensions[get_column_letter(1+j)].width=w

dv=DataValidation(type="list",formula1='"x"',allow_blank=True); ws.add_data_validation(dv)

r=first
for cat,rows in DATA:
    put(r,1,cat,bold=True,fill=NAVY_LT,fc=NAVY,border=True)
    for j in range(1,5): put(r,1+j,"",fill=NAVY_LT,border=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); put(r,1,cat,bold=True,fill=NAVY_LT,fc=NAVY,border=True)
    r+=1
    for feat,desc,typ,h in rows:
        put(r,1,feat,bold=True,wrap=True,border=True)
        put(r,2,desc,wrap=True,size=9,border=True)
        if typ=="P":
            put(r,3,"Pflicht",fill=PFILL,fc="FFFFFF",bold=True,align="center",size=9,border=True)
            put(r,5,"x",align="center",bold=True,border=True)
        else:
            put(r,3,"Optional",fill=OFILL,align="center",size=9,border=True)
            put(r,5,"",align="center",border=True)
        put(r,4,h,align="center",border=True,numfmt='#,##0')
        dv.add(ws.cell(r,5))
        r+=1

ws.freeze_panes="A9"
ws.row_dimensions[HEAD].height=24
wb.save("docs/Feature-Liste-Aufwand.xlsx")
print("ok – Pflicht h:",pflicht_h," Optional h:",opt_h," Datenzeilen bis:",last)
