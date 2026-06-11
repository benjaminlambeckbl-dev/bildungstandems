# -*- coding: utf-8 -*-
"""Erzeugt docs/Feature-Liste.xlsx (Überkategorien, differenziert)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1D1F4E"; NAVY_LT = "E9EAF3"
STAT = {
    "V1": ("V1", "C6EFCE", "1C2230"),       # grün – im Piloten
    "B":  ("V1-Basis", "DDEBF7", "1C2230"),  # blau – Basis, Ausbau folgt
    "A":  ("Ausbau", "F2F2F2", "5C6B7A"),    # grau – Folgebudget
}
thin = Side(style="thin", color="D9DEE6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# (Überkategorie, [(Feature, Beschreibung, status, quelle)])
DATA = [
 ("1 · Terminverwaltung", [
   ("Terminübersicht", "Termine werden allen Beteiligten übersichtlich angezeigt; bei Schüler:innen nur der nächste Termin prominent.", "V1", "P"),
   ("Vorschlagen / Bestätigen / Absagen", "Termine können vorgeschlagen, bestätigt und abgesagt werden; bei Absage automatische Info an die betroffenen Coaches.", "V1", "P"),
   ("Termin-Detail & Doku", "Detailseite mit Teilnahme-Zusage, Vorbereitungs-Checkliste und Treffen-Dokumentation (Anwesenheit).", "V1", "P"),
   ("Kalenderansicht", "Monatskalender mit persönlichen, schulbezogenen und globalen Terminen.", "V1", "P/L"),
   ("Erinnerungen", "Automatischer Hinweis vor einem Treffen über einen Kanal (Push oder E-Mail).", "B", "P/L"),
   ("Schulungs-/Trainingstermine", "Trainingstermine werden von Trainer:innen angelegt; sichtbar für Admin und die betroffene Schule.", "A", "W"),
   ("Termin-Hoheit (Rollen)", "Festlegen, wer Termine einstellt und wer bestätigt (Trainer:in / Lehrkraft / SuS) – Teil des Rollenkonzepts.", "V1", "W"),
   ("Trainings-Lücken-Erkennung", "System erkennt fehlende Trainingstermine je Schule und meldet Lücken aktiv (statt manuell suchen).", "A", "W"),
   ("Terminliste exportieren", "Termine/Trainings als Liste exportierbar (z. B. Excel).", "A", "W"),
   ("Terminabstimmung", "Abstimmung über mögliche Termine (Poll).", "A", "L"),
 ]),
 ("2 · Materialbereitstellung", [
   ("Materialien ansehen", "Liste mit Volltextsuche, Kategorien, Tags und „zuletzt geöffnet“.", "V1", "P/L"),
   ("Auf Plattform lesbar", "Materialien direkt in der App ansehbar, nicht nur als PDF-Download.", "B", "W"),
   ("Lernpfade", "Lern-Module mit Schritten, Quiz und Vorlese-Funktion; Fortschritt wird gespeichert.", "V1", "P"),
   ("FAQ", "Durchsuchbare, kategorisierte Sammlung häufiger Fragen.", "V1", "P/L"),
   ("Zeitgesteuerte Freischaltung", "Inhalte werden ab Datum/Phase sichtbar (vorher gesperrt mit Hinweis).", "B", "P/L/W"),
   ("Auto-Freischaltung nach Training", "Nach 1./2. Training wird automatisch weiteres Begleitmaterial freigeschaltet.", "A", "W"),
   ("Material-CMS mit Editor", "ZSB/Admin legt Materialien selbst an (fett/kursiv/Überschriften), Methodenkarten, Texte, Begleitmaterial.", "A", "W"),
   ("Datei-Upload / Storage", "Hochladen von PDF und Bildern mit Rechteverwaltung.", "A", "L/W"),
   ("Nutzungs-Statistik", "Wie oft wurde ein Material geöffnet (aggregiert; personenbezogen voraussichtlich nicht nötig).", "A", "W"),
 ]),
 ("3 · Evaluation, Feedback & Dokumentation", [
   ("Reflexion nach Treffen", "Coach hält Stimmung und Freitext fest (was lief gut / schwierig).", "V1", "P"),
   ("Tägliches Kurz-Feedback", "Niedrigschwelliger Stimmungscheck „Wie lief es?“ per Button.", "V1", "P/L"),
   ("„Hilfe nötig?“ (Coach)", "Coach fordert per Knopf Unterstützung an; Koordination wird informiert.", "V1", "P"),
   ("„Hilfe nötig?“ (Koordination)", "Auch Schulkoordination kann niedrigschwellig „läuft nicht rund“ an die ZSB melden.", "B", "W"),
   ("Treffen dokumentieren", "Durchgeführt ja/nein, Anzahl anwesender Kinder, kurze Notiz.", "V1", "P"),
   ("Rückmeldung an Coach", "Koordination gibt individuelle Rückmeldung/Lob (Freitext) an den Coach.", "V1", "P"),
   ("Änderungs-Verlauf / Historie", "Korrekturen werden nicht überschrieben; Verlauf sichtbar (z. B. Gruppe 15→5, Drop-outs erkennbar).", "A", "W"),
 ]),
 ("4 · Auswertung & Monitoring", [
   ("Kennzahlen ZSB", "Erreichte Kinder, Anzahl Tandems, Termine etc. in der ZSB-Ansicht.", "V1", "P"),
   ("Diagramme", "Visualisierung von Stimmung/Feedback je Schule und Zeitraum.", "B", "P/L"),
   ("Ampel-Monitoring", "Schulen werden grün/gelb/rot eingefärbt (ggf. an Stimmung gekoppelt); Anzahl je Status sichtbar.", "A", "W"),
   ("Frühwarnung", "Schulen mit Schwierigkeiten werden hervorgehoben → gezielte Begleitung/Intervention.", "A", "L/W"),
   ("Auswertung exportieren", "Stimmung/Kennzahlen je Schule exportierbar.", "A", "W"),
 ]),
 ("5 · Kommunikation", [
   ("Info-Kanal Lehrkraft→Coach", "Einseitige Infos/Absagen an Coaches.", "V1", "P"),
   ("Lehrkraft ↔ Grundschul-Koordination", "Direktnachricht auf der organisatorischen Ebene zwischen den kooperierenden Schulen.", "B", "W"),
   ("Gruppenchat", "Koordination kann Tandems/Coach-Gruppen gemeinsam anschreiben (Coaches arbeiten zu zweit/dritt).", "A", "L/W"),
   ("Schul-/Partnerschulgruppen", "Automatisch generierte Gruppen je Schule/Partnerschaft.", "A", "L"),
   ("Wichtige Nachricht „Sticky“", "Hervorheben/Anheften wichtiger Nachrichten.", "A", "L"),
   ("Moderation / Meldefunktion", "Melde-/Moderationswerkzeuge (Kinderschutz) für Chats unter Minderjährigen.", "A", "L"),
 ]),
 ("6 · Rollen, Rechte & Nutzerverwaltung", [
   ("Rollen", "Admin (ZSB), Trainer:in (ZSB), Koordination (Schule), Schüler:in/Coach.", "V1", "P/L/W"),
   ("Rollen-/Rechtematrix (RLS)", "Serverseitig durchgesetzt: wer darf sehen/hochladen/bearbeiten und wer wem schreiben.", "V1", "L/W"),
   ("Trainer:in-Rolle", "Eigene Rolle; sieht nur die betreuten Schulen, nicht die ZSB-Meta-Ebene.", "A", "W"),
   ("Schulen anlegen / Nutzer einladen", "Admin legt Schulen an, lädt Nutzer ein und ordnet sie zu.", "B", "L/W"),
   ("Trainer zuweisen", "Trainer:innen werden Schulen zugewiesen (nachträglich änderbar).", "A", "L/W"),
   ("Jahreswechsel-Logik", "Jährliche Veränderungen der Nutzerstruktur abbilden (Übergang/Archiv).", "A", "L"),
   ("Admin-Backend", "ZSB stellt Termine und Materialien für Schulen selbst ein.", "B", "W"),
 ]),
 ("7 · Datenmodell & Schulstruktur", [
   ("Schul-Kooperation", "Grundschule ↔ weiterführende Schule verbunden; eine weiterführende kann bis zu 3 Grundschulen betreuen.", "V1", "W"),
   ("Tandem-Definition", "Tandem = Coach (Jugendliche:r) + Coachees (Grundschulkinder). Grundschulkinder nur mit Vornamen, NICHT in der App.", "V1", "P/W"),
   ("Tandems unter Schule", "In der ZSB-Ansicht keinen eigenen „Tandems“-Reiter; Tandems unter der jeweiligen Schule.", "V1", "W"),
   ("Schul-Filter / Suche", "Bei >100 Schulen: Filter/Suche nach Stadt, Schulform, Sozialindex.", "A", "W"),
 ]),
 ("8 · Onboarding & Login", [
   ("Individueller Login", "Sichere, individuelle Anmeldung für SuS und Lehrkräfte.", "V1", "L/W"),
   ("Einladungs-/Onboarding-Flow", "Geführter Einstieg für neue Nutzer:innen.", "V1", "P/W"),
   ("QR-Code-Onboarding", "Schule anlegen → QR-Code → SuS scannen → automatisch der Schule zugeordnet (kein Einzel-Eintippen).", "A", "W"),
   ("App-Einführung in Schulungen", "Die App wird bereits in den Coach-Schulungen eingeführt/genutzt.", "A", "W"),
 ]),
 ("9 · Export & Berichte", [
   ("Coach-Zertifikat", "Druckbare Urkunde (PDF); Namen der SuS werden erfasst (Tippfehler-Risiko bedenken).", "V1", "P/W"),
   ("Excel-/Berichts-Export", "Export von durchgeführten Trainings, Schülerzahlen, Stadtverteilung, Stimmung und Namen.", "A", "W"),
 ]),
 ("10 · Nicht-funktional & Design", [
   ("PWA / Offline", "Wie eine App installierbar; Inhalte auch offline lesbar.", "V1", "P"),
   ("Responsive Design", "Optimiert für Handy und Desktop.", "V1", "P/W"),
   ("Schlankes Schüler-UI", "Minimalistischer, weniger Emojis/Piktogramme, weniger Scrollen.", "V1", "W"),
   ("DSGVO-Basis", "EU-Hosting, Datensparsamkeit, Einwilligungen, rollenbasierter Zugriff.", "B", "L/W"),
   ("Performance", "Stabile Nutzung bei >1000 Nutzer:innen, schnelle Ladezeiten.", "A", "L"),
   ("CI-/Oberflächen-Anpassbarkeit", "Farben/Design anpassbar an das Corporate Design.", "A", "W"),
 ]),
]

wb = Workbook(); ws = wb.active; ws.title = "Feature-Liste"
ws.sheet_view.showGridLines = False

def put(r, c, v, *, bold=False, fill=None, fc="1C2230", size=10, align="left", wrap=False, border=False):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = Font(name="Calibri", size=size, bold=bold, color=fc)
    cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill: cl.fill = PatternFill("solid", fgColor=fill)
    if border: cl.border = BORDER
    return cl

put(1, 1, "Feature-Liste – BildungsTandems / ZSB", bold=True, size=15, fc=NAVY)
put(2, 1, "Differenziert nach Überkategorien · Prototyp + Lastenheft + Workshop (11.06.2026) · Coachingspace GmbH", size=9, fc="5C6B7A")
put(3, 1, "Status:", bold=True, size=9, fc=NAVY)
for i, k in enumerate(["V1", "B", "A"]):
    t, bg, fcc = STAT[k]
    put(3, 2 + i, {"V1":"V1 = im 20k-Piloten","B":"V1-Basis (Ausbau folgt)","A":"Ausbau (Folgebudget)"}[k], fill=bg, fc=fcc, align="center", size=9, bold=True, border=True)
put(3, 5, "Quelle: P=Prototyp · L=Lastenheft · W=Workshop", size=9, fc="5C6B7A")

HEAD = 5
heads = [("Feature", 30), ("Beschreibung (wer sieht / darf was)", 78), ("Status", 12), ("Quelle", 10)]
for j, (h, w) in enumerate(heads):
    put(HEAD, 1 + j, h, bold=True, fill=NAVY, fc="FFFFFF", align="left", border=True)
    ws.column_dimensions[get_column_letter(1 + j)].width = w

r = HEAD + 1
for cat, rows in DATA:
    put(r, 1, cat, bold=True, fill=NAVY_LT, fc=NAVY, border=True)
    for j in range(1, 4):
        put(r, 1 + j, "", fill=NAVY_LT, border=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    put(r, 1, cat, bold=True, fill=NAVY_LT, fc=NAVY, border=True)
    r += 1
    for feat, desc, st, q in rows:
        t, bg, fcc = STAT[st]
        put(r, 1, feat, bold=True, wrap=True, border=True)
        put(r, 2, desc, wrap=True, border=True, size=9)
        put(r, 3, t, fill=bg, fc=fcc, align="center", bold=True, border=True)
        put(r, 4, q, align="center", size=9, fc="5C6B7A", border=True)
        r += 1

ws.freeze_panes = "A6"
ws.row_dimensions[HEAD].height = 26
wb.save("docs/Feature-Liste.xlsx")
print("xlsx erstellt, Zeilen:", r)
